from pathlib import Path

from openpyxl import Workbook

from hotel_supply_demand.fetcher import sha256_file
from hotel_supply_demand.municipality.database import load_municipality_records
from hotel_supply_demand.municipality.parser import parse_municipality_workbook
from hotel_supply_demand.municipality.sources import MunicipalitySource


def make_municipality_workbook(path: Path, *, omit_table: int | None = None) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    locations = ["北海道札幌市", "沖縄県石垣市"]
    for table in (5, 6, 8, 11, 12):
        if table == omit_table:
            continue
        sheet = workbook.create_sheet(f"参考第{table}表(5月)")
        sheet.cell(1, 1, f"参考第{table}表　施設所在地(主な市区町村)のテスト")
        sheet.cell(4, 1, "施設所在地\n（主な市区町村）")
        for offset, location in enumerate(locations, 7):
            sheet.cell(offset, 1, location)
        if table == 5:
            values = [[2, "-", 3, "-", 4, 4], [5, 5, 2, "-", 10, 8]]
        else:
            values = {
                6: [[100, "-", "-", 100], [200, 20, "-", 180]],
                8: [[25, "-", "-", 25], [60, 5, "-", 55]],
                11: [[50, "-", "-", 50], [80, 8, "-", 72]],
                12: [[60.5, "-", "-", 60.5], [70.0, 20.0, "-", 72.0]],
            }[table]
        for row, row_values in zip(range(7, 9), values, strict=True):
            for column, value in enumerate(row_values, 2):
                sheet.cell(row, column, value)
    workbook.save(path)


def make_municipality_report_database(root: Path) -> Path:
    workbook = root / "monthly.xlsx"
    database = root / "hotel_market.sqlite3"
    make_municipality_workbook(workbook)
    records = parse_municipality_workbook(workbook, 2026, 5)
    source = MunicipalitySource(
        year=2026,
        month=5,
        release_type="second_preliminary",
        stat_inf_id="000040482630",
        published_on="2026-07-31",
        url="https://www.e-stat.go.jp/stat-search/file-download?statInfId=000040482630&fileKind=0",
        filename="2026-05_second_preliminary.xlsx",
    )
    load_municipality_records(
        database,
        records,
        source,
        {
            "retrieved_at": "2026-08-15T00:00:00+00:00",
            "sha256": sha256_file(workbook),
            "size_bytes": workbook.stat().st_size,
        },
    )
    return database
