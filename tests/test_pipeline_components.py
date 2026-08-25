import io
import json
import sqlite3
import tempfile
import unittest
from pathlib import Path
from typing import Self
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from hotel_supply_demand.fetcher import FetchError, sha256_file, validate_xlsx
from hotel_supply_demand.prefecture.database import (
    build_database,
    migrate_legacy_prefecture_market,
)
from hotel_supply_demand.prefecture.fetcher import fetch_sources
from hotel_supply_demand.prefecture.parser import (
    PREFECTURE_NAMES,
    parse_national_occupancy,
    parse_workbook,
)
from hotel_supply_demand.prefecture.sources import Source
from hotel_supply_demand.prefecture.validation import DataQualityError, validate_records


def make_workbook(path: Path, year: int = 2025) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for month in range(1, 13):
        for table in (1, 4, 8):
            sheet = workbook.create_sheet(f"第{table}表({month}月)")
            sheet["A9"] = f"令和{year - 2018}年{month}月"
            if table == 1:
                sheet["C4"] = "総数\n1)、2)"
            elif table == 4:
                sheet["C4"] = "延べ\n宿泊者数\n1)"
                sheet["J4"] = "うち\n外国人延べ\n宿泊者数\n1)"
            else:
                sheet["C4"] = "客室稼働率\n1)、2)"
                sheet["C9"] = 55
            for code in range(1, 48):
                row = 9 + code
                sheet.cell(row, 1, f"{code:02d}地域{code}県")
                sheet.cell(row, 3, 50 if table == 8 else 100 + code)
                if table == 4:
                    sheet.cell(row, 10, 10 + code)
    workbook.save(path)


class FakeDownload(io.BytesIO):
    class Headers:
        @staticmethod
        def get_content_type() -> str:
            return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    headers = Headers()

    def __enter__(self) -> Self:
        return self

    def __exit__(self, *_: object) -> None:
        self.close()


class PipelineComponentsTest(unittest.TestCase):
    def test_migrate_legacy_prefecture_demand_and_supply(self):
        connection = sqlite3.connect(":memory:")
        try:
            connection.executescript(
                """CREATE TABLE source_files(id INTEGER PRIMARY KEY);
                   CREATE TABLE prefectures(code INTEGER PRIMARY KEY,name TEXT);
                   CREATE TABLE monthly_demand(
                     year,month,prefecture_code,release_type,total_guests,
                     japanese_guests,foreign_guests,occupancy_rate,source_file_id
                   );
                   CREATE TABLE monthly_supply(
                     year,month,prefecture_code,release_type,facilities,source_file_id
                   );
                   INSERT INTO source_files VALUES(1);
                   INSERT INTO prefectures VALUES(1,'北海道');
                   INSERT INTO monthly_demand
                     VALUES(2025,1,1,'final',100,80,20,60.0,1);
                   INSERT INTO monthly_supply VALUES(2025,1,1,'final',10,1);
                   CREATE VIEW latest_monthly_demand AS SELECT * FROM monthly_demand;
                   CREATE VIEW latest_monthly_supply AS SELECT * FROM monthly_supply;
                   CREATE VIEW monthly_market AS SELECT 1;
                   CREATE VIEW annual_market AS SELECT 1;"""
            )
            self.assertTrue(migrate_legacy_prefecture_market(connection))
            row = connection.execute(
                "SELECT total_guests,facilities,prefecture_name FROM latest_prefecture_market"
            ).fetchone()
            self.assertEqual(row, (100, 10, "北海道"))
            names = {
                item[0]
                for item in connection.execute("SELECT name FROM sqlite_master WHERE type='table'")
            }
            self.assertNotIn("monthly_demand", names)
            self.assertNotIn("monthly_supply", names)
        finally:
            connection.close()

    def test_parse_estat_workbook_without_codes_or_period_caption(self):
        with tempfile.TemporaryDirectory() as directory:
            xlsx = Path(directory) / "estat.xlsx"
            make_workbook(xlsx)
            workbook = load_workbook(xlsx)
            for month in range(1, 13):
                for table in (1, 4, 8):
                    sheet = workbook[f"第{table}表({month}月)"]
                    sheet["A9"] = "全国" if table == 8 else "施設所在地計"
                    for code, name in enumerate(PREFECTURE_NAMES, 1):
                        sheet.cell(9 + code, 1, name)
            workbook.save(xlsx)

            records = parse_workbook(xlsx, 2025)
            national = parse_national_occupancy(xlsx, 2025)

            self.assertEqual(len(records), 564)
            self.assertEqual(records[0].prefecture_name, "北海道")
            self.assertEqual(len(national), 12)
            self.assertEqual(national[0].occupancy_rate, 55)

    def test_fetch_replaces_stale_destination_even_when_download_hash_exists_elsewhere(self):
        with tempfile.TemporaryDirectory() as directory:
            raw_dir = Path(directory)
            destination = raw_dir / "2025_final.xlsx"
            download = raw_dir / "download.xlsx"
            make_workbook(destination, 2024)
            make_workbook(download, 2025)
            downloaded_bytes = download.read_bytes()
            downloaded_hash = sha256_file(download)
            manifest = {
                "schema_version": 1,
                "files": [
                    {
                        "year": 2024,
                        "release_type": "final",
                        "url": "https://www.mlit.go.jp/2024.xlsx",
                        "filename": "2024_final.xlsx",
                        "sha256": downloaded_hash,
                    },
                    {
                        "year": 2025,
                        "release_type": "final",
                        "url": "https://www.mlit.go.jp/2025.xlsx",
                        "filename": destination.name,
                        "sha256": "0" * 64,
                    },
                ],
            }
            (raw_dir / "manifest.json").write_text(json.dumps(manifest), encoding="utf-8")
            source = Source(
                year=2025,
                release_type="final",
                url="https://www.mlit.go.jp/2025.xlsx",
                filename=destination.name,
                published_on="2026-07-06",
            )

            with patch(
                "hotel_supply_demand.prefecture.fetcher.urlopen",
                return_value=FakeDownload(downloaded_bytes),
            ):
                result = fetch_sources([source], raw_dir)

            updated_manifest = json.loads((raw_dir / "manifest.json").read_text(encoding="utf-8"))
            updated_entry = next(item for item in updated_manifest["files"] if item["year"] == 2025)
            self.assertEqual(result[0]["status"], "downloaded")
            self.assertEqual(sha256_file(destination), downloaded_hash)
            self.assertEqual(updated_entry["sha256"], downloaded_hash)
            self.assertEqual(updated_entry["revisions"][0]["sha256"], "0" * 64)

    def test_parse_validate_and_load(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            xlsx = root / "source.xlsx"
            database = root / "result.sqlite3"
            make_workbook(xlsx)
            records = parse_workbook(xlsx, 2025)
            national = parse_national_occupancy(xlsx, 2025)
            self.assertEqual(validate_records(records, {2025})["rows"], 564)
            manifest = [
                {
                    "year": 2025,
                    "release_type": "final",
                    "url": "https://www.mlit.go.jp/source.xlsx",
                    "filename": "source.xlsx",
                    "published_on": "2026-07-06",
                    "retrieved_at": "2026-01-01T00:00:00+00:00",
                    "sha256": sha256_file(xlsx),
                    "size_bytes": xlsx.stat().st_size,
                }
            ]
            build_database(database, records, manifest, national)
            connection = sqlite3.connect(database)
            try:
                self.assertEqual(
                    connection.execute("SELECT count(*) FROM monthly_prefecture_market").fetchone()[
                        0
                    ],
                    564,
                )
                self.assertEqual(
                    connection.execute("SELECT count(*) FROM latest_prefecture_market").fetchone()[
                        0
                    ],
                    564,
                )
                self.assertEqual(
                    connection.execute("SELECT count(*) FROM national_occupancy").fetchone()[0], 12
                )
            finally:
                connection.close()

    def test_invalid_xlsx_and_incomplete_records_fail(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "bad.xlsx"
            path.write_text("not an xlsx")
            with self.assertRaises(FetchError):
                validate_xlsx(path)
        with self.assertRaises(DataQualityError):
            validate_records([], {2025})
