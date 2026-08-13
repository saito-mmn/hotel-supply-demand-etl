"""Parse annual final-value workbooks into a stable monthly schema."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from .models import MonthlyRecord


class WorkbookFormatError(ValueError):
    pass


PREFECTURE = re.compile(r"^\s*(\d{2})(.+?[都道府県])\s*$")
JAPANESE_DATE = re.compile(r"(平成|令和)(元|\d+)年(\d+)月")


def _number(value: object) -> int | float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "")
        if cleaned in {"", "-", "…", "...", "X", "x", "*"}:
            return None
        try:
            return float(cleaned) if "." in cleaned else int(cleaned)
        except ValueError:
            return None
    return None


def _sheet_values(workbook, title: str, column: int) -> dict[int, tuple[str, int | float | None]]:
    if title not in workbook.sheetnames:
        raise WorkbookFormatError(f"missing worksheet: {title}")
    sheet = workbook[title]
    result = {}
    for row in sheet.iter_rows(min_row=7, values_only=True):
        match = PREFECTURE.match(str(row[0] or ""))
        if not match:
            continue
        code = int(match.group(1))
        if 1 <= code <= 47:
            result[code] = (match.group(2).strip(), _number(row[column - 1]))
    if set(result) != set(range(1, 48)):
        missing = sorted(set(range(1, 48)) - set(result))
        raise WorkbookFormatError(f"{title}: prefectures missing: {missing}")
    return result


def _validate_period(workbook, title: str, expected_year: int, expected_month: int) -> None:
    value = str(workbook[title]["A7"].value or "")
    match = JAPANESE_DATE.search(value)
    if not match:
        raise WorkbookFormatError(f"{title}: cannot identify reporting period from A7")
    era_year = 1 if match.group(2) == "元" else int(match.group(2))
    actual_year = (1988 if match.group(1) == "平成" else 2018) + era_year
    actual_month = int(match.group(3))
    if (actual_year, actual_month) != (expected_year, expected_month):
        raise WorkbookFormatError(
            f"{title}: period {actual_year}-{actual_month:02} != {expected_year}-{expected_month:02}"
        )


def parse_workbook(path: Path, year: int, release_type: str = "final") -> list[MonthlyRecord]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[MonthlyRecord] = []
    try:
        for month in range(1, 13):
            _validate_period(workbook, f"第1表({month}月)", year, month)
            _validate_period(workbook, f"第4表({month}月)", year, month)
            _validate_period(workbook, f"第8表({month}月)", year, month)
            facilities = _sheet_values(workbook, f"第1表({month}月)", 2)
            total = _sheet_values(workbook, f"第4表({month}月)", 2)
            foreign = _sheet_values(workbook, f"第4表({month}月)", 9)
            occupancy = _sheet_values(workbook, f"第8表({month}月)", 2)
            for code in range(1, 48):
                names = {facilities[code][0], total[code][0], foreign[code][0], occupancy[code][0]}
                if len(names) != 1:
                    raise WorkbookFormatError(f"inconsistent prefecture name: {year}-{month:02} {code}")
                total_value = total[code][1]
                foreign_value = foreign[code][1]
                japanese = None
                if total_value is not None and foreign_value is not None:
                    japanese = int(total_value) - int(foreign_value)
                records.append(
                    MonthlyRecord(
                        year=year,
                        month=month,
                        prefecture_code=code,
                        prefecture_name=names.pop(),
                        total_guests=int(total_value) if total_value is not None else None,
                        foreign_guests=int(foreign_value) if foreign_value is not None else None,
                        japanese_guests=japanese,
                        occupancy_rate=float(occupancy[code][1]) if occupancy[code][1] is not None else None,
                        facilities=int(facilities[code][1]) if facilities[code][1] is not None else None,
                        release_type=release_type,
                    )
                )
    finally:
        workbook.close()
    return records
