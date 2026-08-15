"""Parse municipality-level reference tables from monthly second-preliminary XLSX files."""

from __future__ import annotations

from pathlib import Path
from typing import TypeAlias

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import MunicipalityMonthlyRecord


class MunicipalityWorkbookFormatError(ValueError):
    pass


PREFECTURES = (
    "北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県",
    "茨城県", "栃木県", "群馬県", "埼玉県", "千葉県", "東京都", "神奈川県",
    "新潟県", "富山県", "石川県", "福井県", "山梨県", "長野県", "岐阜県",
    "静岡県", "愛知県", "三重県", "滋賀県", "京都府", "大阪府", "兵庫県",
    "奈良県", "和歌山県", "鳥取県", "島根県", "岡山県", "広島県", "山口県",
    "徳島県", "香川県", "愛媛県", "高知県", "福岡県", "佐賀県", "長崎県",
    "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県",
)
PREFECTURE_CODES = {name: code for code, name in enumerate(PREFECTURES, 1)}
ROOM_SIZE_COLUMNS = {"total": 2, "1_to_9": 3, "10_to_19": 4, "20_plus": 5}
MISSING_VALUES = {"", "-", "－", "…", "...", "X", "x", "*"}
FULL_WIDTH_TABLE_NUMBERS = {5: "５", 6: "６", 8: "８", 11: "１１", 12: "１２"}

MetricKey: TypeAlias = tuple[str, str]


def _number(value: object, *, integer: bool) -> int | float | None:
    if value is None or str(value).strip() in MISSING_VALUES:
        return None
    try:
        number = float(str(value).replace(",", "").strip())
    except ValueError as exc:
        raise MunicipalityWorkbookFormatError(f"unexpected numeric value: {value!r}") from exc
    if integer:
        if not number.is_integer():
            raise MunicipalityWorkbookFormatError(f"expected integer value: {value!r}")
        return int(number)
    return number


def _location(value: object) -> tuple[int, str, str] | None:
    raw = str(value or "").strip().replace(" ", "").replace("　", "")
    for prefecture in PREFECTURES:
        if raw.startswith(prefecture) and len(raw) > len(prefecture):
            municipality = raw[len(prefecture) :]
            return PREFECTURE_CODES[prefecture], prefecture, municipality
    return None


def _sheet(workbook, table_number: int, month: int) -> Worksheet:
    name = f"参考第{table_number}表({month}月)"
    if name not in workbook.sheetnames:
        raise MunicipalityWorkbookFormatError(f"missing worksheet: {name}")
    sheet = workbook[name]
    title = str(sheet.cell(1, 1).value or "")
    expected_titles = {
        f"参考第{table_number}表",
        f"参考第{FULL_WIDTH_TABLE_NUMBERS[table_number]}表",
    }
    if not any(expected_title in title for expected_title in expected_titles):
        raise MunicipalityWorkbookFormatError(f"unexpected worksheet title: {name}")
    if not any(
        "主な市区町村" in str(sheet.cell(row, 1).value or "")
        for row in range(4, 7)
    ):
        raise MunicipalityWorkbookFormatError(f"municipality header not found: {name}")
    return sheet


def _metric_values(
    sheet: Worksheet, *, metric_name: str, integer: bool
) -> tuple[dict[MetricKey, int | float | None], dict[str, tuple[int, str, str]]]:
    values: dict[MetricKey, int | float | None] = {}
    locations: dict[str, tuple[int, str, str]] = {}
    for row in range(7, sheet.max_row + 1):
        location = _location(sheet.cell(row, 1).value)
        if location is None:
            continue
        raw_name = str(sheet.cell(row, 1).value).strip().replace(" ", "").replace("　", "")
        if raw_name in locations:
            raise MunicipalityWorkbookFormatError(
                f"duplicate municipality in {sheet.title}: {raw_name}"
            )
        locations[raw_name] = location
        for room_size, column in ROOM_SIZE_COLUMNS.items():
            values[(raw_name, room_size)] = _number(
                sheet.cell(row, column).value, integer=integer
            )
    if not locations:
        raise MunicipalityWorkbookFormatError(
            f"no municipality rows found for {metric_name}: {sheet.title}"
        )
    return values, locations


def _sum_complete(values: list[int | None]) -> int | None:
    return sum(values) if all(value is not None for value in values) else None


def _facility_values(
    sheet: Worksheet,
) -> tuple[dict[MetricKey, tuple[int | None, int | None]], dict[str, tuple[int, str, str]]]:
    values: dict[MetricKey, tuple[int | None, int | None]] = {}
    locations: dict[str, tuple[int, str, str]] = {}
    columns = {"1_to_9": (2, 3), "10_to_19": (4, 5), "20_plus": (6, 7)}
    for row in range(7, sheet.max_row + 1):
        location = _location(sheet.cell(row, 1).value)
        if location is None:
            continue
        raw_name = str(sheet.cell(row, 1).value).strip().replace(" ", "").replace("　", "")
        if raw_name in locations:
            raise MunicipalityWorkbookFormatError(
                f"duplicate municipality in {sheet.title}: {raw_name}"
            )
        locations[raw_name] = location
        populations: list[int | None] = []
        responses: list[int | None] = []
        for room_size, (population_column, response_column) in columns.items():
            population = _number(sheet.cell(row, population_column).value, integer=True)
            response = _number(sheet.cell(row, response_column).value, integer=True)
            populations.append(population if isinstance(population, int) else None)
            responses.append(response if isinstance(response, int) else None)
            values[(raw_name, room_size)] = (
                population if isinstance(population, int) else None,
                response if isinstance(response, int) else None,
            )
        values[(raw_name, "total")] = (
            _sum_complete(populations),
            _sum_complete(responses),
        )
    if not locations:
        raise MunicipalityWorkbookFormatError(
            f"no municipality rows found for facilities: {sheet.title}"
        )
    return values, locations


def parse_municipality_workbook(
    path: Path,
    year: int,
    month: int,
    release_type: str = "second_preliminary",
) -> list[MunicipalityMonthlyRecord]:
    """Join municipality reference tables into one record per municipality and room class."""
    if not 1 <= month <= 12:
        raise MunicipalityWorkbookFormatError(f"invalid month: {month}")
    if release_type != "second_preliminary":
        raise MunicipalityWorkbookFormatError(f"unsupported release type: {release_type}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        facilities, facility_locations = _facility_values(_sheet(workbook, 5, month))
        total_guests, locations = _metric_values(
            _sheet(workbook, 6, month), metric_name="total guests", integer=True
        )
        foreign_guests, foreign_locations = _metric_values(
            _sheet(workbook, 8, month), metric_name="foreign guests", integer=True
        )
        occupied_rooms, room_locations = _metric_values(
            _sheet(workbook, 11, month), metric_name="occupied rooms", integer=False
        )
        occupancy_rate, occupancy_locations = _metric_values(
            _sheet(workbook, 12, month), metric_name="occupancy rate", integer=False
        )
    finally:
        workbook.close()

    all_locations: dict[str, tuple[int, str, str]] = {}
    for table_locations in (
        facility_locations,
        locations,
        foreign_locations,
        room_locations,
        occupancy_locations,
    ):
        for raw_name, location in table_locations.items():
            previous = all_locations.setdefault(raw_name, location)
            if previous != location:
                raise MunicipalityWorkbookFormatError(
                    f"inconsistent municipality name: {raw_name}"
                )

    records: list[MunicipalityMonthlyRecord] = []
    for raw_name in all_locations:
        prefecture_code, prefecture_name, municipality_name = all_locations[raw_name]
        for room_size in ROOM_SIZE_COLUMNS:
            key = (raw_name, room_size)
            total = total_guests.get(key)
            foreign = foreign_guests.get(key)
            occupancy = occupancy_rate.get(key)
            if total is not None and foreign is not None and foreign > total:
                raise MunicipalityWorkbookFormatError(
                    f"foreign guests exceed total guests: {raw_name} {room_size}"
                )
            if occupancy is not None and not 0 <= occupancy <= 200:
                raise MunicipalityWorkbookFormatError(
                    f"occupancy rate out of range: {raw_name} {room_size} {occupancy}"
                )
            population, responding = facilities.get(key, (None, None))
            records.append(
                MunicipalityMonthlyRecord(
                    year=year,
                    month=month,
                    prefecture_code=prefecture_code,
                    prefecture_name=prefecture_name,
                    municipality_name=municipality_name,
                    total_guests=total if isinstance(total, int) else None,
                    japanese_guests=(total - foreign)
                    if isinstance(total, int) and isinstance(foreign, int)
                    else None,
                    foreign_guests=foreign if isinstance(foreign, int) else None,
                    occupied_rooms=float(occupied_rooms[key])
                    if occupied_rooms.get(key) is not None
                    else None,
                    occupancy_rate=float(occupancy) if occupancy is not None else None,
                    population_facilities=population,
                    responding_facilities=responding,
                    room_size_class=room_size,
                    release_type=release_type,
                )
            )
    return records
