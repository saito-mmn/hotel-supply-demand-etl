"""Parse annual final-value workbooks into a stable monthly schema."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from .models import MonthlyRecord, NationalOccupancyRecord


class WorkbookFormatError(ValueError):
    pass


PREFECTURE_NAMES = (
    "北海道", "青森県", "岩手県", "宮城県", "秋田県", "山形県", "福島県",
    "茨城県", "栃木県", "群馬県", "埼玉県", "千葉県", "東京都", "神奈川県",
    "新潟県", "富山県", "石川県", "福井県", "山梨県", "長野県", "岐阜県",
    "静岡県", "愛知県", "三重県", "滋賀県", "京都府", "大阪府", "兵庫県",
    "奈良県", "和歌山県", "鳥取県", "島根県", "岡山県", "広島県", "山口県",
    "徳島県", "香川県", "愛媛県", "高知県", "福岡県", "佐賀県", "長崎県",
    "熊本県", "大分県", "宮崎県", "鹿児島県", "沖縄県",
)
PREFECTURE_CODES = {name: code for code, name in enumerate(PREFECTURE_NAMES, 1)}
PREFECTURE = re.compile(r"^\s*(?:(\d{2})\s*)?(.+?[都道府県])\s*$")
JAPANESE_DATE = re.compile(r"(平成|令和)(元|\d+)年(\d+)月")
HEADER_SCAN_ROWS = 15

# Known schema of the official 2019-2025 annual final-value workbooks. The table
# numbers are stable across the supported files, but are not assumed to be a
# permanent government API contract. Sheet, period, header, and prefecture
# validations below intentionally stop the ETL if a future workbook changes.
FINAL_WORKBOOK_METRICS = {
    "facilities": ("第1表", "総数"),
    "total_guests": ("第4表", "延べ宿泊者数"),
    "foreign_guests": ("第4表", "うち外国人延べ宿泊者数"),
    "occupancy_rate": ("第8表", "客室稼働率"),
}


def _number(value: object) -> int | float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "")
        if cleaned in {"", "-", "…", "...", "X", "x", "*"}:
            return None
        try:
            return float(cleaned) if "." in cleaned else int(cleaned)
        except ValueError:
            return None
    return None


def _normalized_text(value: object) -> str:
    return re.sub(r"\s+", "", str(value or ""))


def _find_column(workbook, title: str, header_prefix: str) -> int:
    sheet = workbook[title]
    matches: set[int] = set()
    for row in sheet.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS, values_only=True):
        for column, value in enumerate(row, 1):
            if _normalized_text(value).startswith(header_prefix):
                matches.add(column)
    if len(matches) != 1:
        raise WorkbookFormatError(
            f"{title}: expected one column starting with {header_prefix!r}, found {sorted(matches)}"
        )
    return matches.pop()


def _sheet_values(
    workbook, title: str, header_prefix: str
) -> dict[int, tuple[str, int | float | None]]:
    if title not in workbook.sheetnames:
        raise WorkbookFormatError(f"missing worksheet: {title}")
    sheet = workbook[title]
    column = _find_column(workbook, title, header_prefix)
    result = {}
    for row in sheet.iter_rows(values_only=True):
        match = PREFECTURE.match(str(row[0] or ""))
        if not match:
            continue
        name = match.group(2).strip()
        code = int(match.group(1)) if match.group(1) else PREFECTURE_CODES.get(name, 0)
        if 1 <= code <= 47:
            result[code] = (name, _number(row[column - 1]))
    if set(result) != set(range(1, 48)):
        missing = sorted(set(range(1, 48)) - set(result))
        raise WorkbookFormatError(f"{title}: prefectures missing: {missing}")
    return result


def _validate_period(workbook, title: str, expected_year: int, expected_month: int) -> None:
    sheet = workbook[title]
    matches = []
    for row in sheet.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS, values_only=True):
        for value in row:
            match = JAPANESE_DATE.search(str(value or ""))
            if match:
                matches.append(match)
    # Newer e-Stat original workbooks omit the reporting-period caption from
    # some/all tables. In that format the sheet name still fixes the month and
    # validated source metadata fixes the year; validate any caption that is
    # present, but do not invent one when the official workbook omits it.
    if not matches:
        return
    if len(matches) != 1:
        raise WorkbookFormatError(
            f"{title}: expected one reporting period in the first {HEADER_SCAN_ROWS} rows, found {len(matches)}"
        )
    match = matches[0]
    era_year = 1 if match.group(2) == "元" else int(match.group(2))
    actual_year = (1988 if match.group(1) == "平成" else 2018) + era_year
    actual_month = int(match.group(3))
    if (actual_year, actual_month) != (expected_year, expected_month):
        raise WorkbookFormatError(
            f"{title}: period {actual_year}-{actual_month:02} != {expected_year}-{expected_month:02}"
        )


def parse_workbook(path: Path, year: int, release_type: str = "final") -> list[MonthlyRecord]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    records: list[MonthlyRecord] = []
    try:
        for month in range(1, 13):
            sheet_titles = {
                table: f"{table}({month}月)"
                for table, _ in FINAL_WORKBOOK_METRICS.values()
            }
            for title in sheet_titles.values():
                _validate_period(workbook, title, year, month)
            values = {
                metric: _sheet_values(workbook, sheet_titles[table], header)
                for metric, (table, header) in FINAL_WORKBOOK_METRICS.items()
            }
            facilities = values["facilities"]
            total = values["total_guests"]
            foreign = values["foreign_guests"]
            occupancy = values["occupancy_rate"]
            for code in range(1, 48):
                names = {facilities[code][0], total[code][0], foreign[code][0], occupancy[code][0]}
                if len(names) != 1:
                    raise WorkbookFormatError(f"inconsistent prefecture name: {year}-{month:02} {code}")
                total_value = total[code][1]
                foreign_value = foreign[code][1]
                japanese = None
                if total_value is not None and foreign_value is not None:
                    japanese = int(total_value) - int(foreign_value)
                records.append(
                    MonthlyRecord(
                        year=year,
                        month=month,
                        prefecture_code=code,
                        prefecture_name=names.pop(),
                        total_guests=int(total_value) if total_value is not None else None,
                        foreign_guests=int(foreign_value) if foreign_value is not None else None,
                        japanese_guests=japanese,
                        occupancy_rate=float(occupancy[code][1]) if occupancy[code][1] is not None else None,
                        facilities=int(facilities[code][1]) if facilities[code][1] is not None else None,
                        release_type=release_type,
                    )
                )
    finally:
        workbook.close()
    return records


def parse_national_occupancy(
    path: Path, year: int, release_type: str = "final"
) -> list[NationalOccupancyRecord]:
    """Read the official national occupancy value from each monthly sheet."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    records = []
    table, header = FINAL_WORKBOOK_METRICS["occupancy_rate"]
    try:
        for month in range(1, 13):
            title = f"{table}({month}月)"
            _validate_period(workbook, title, year, month)
            sheet = workbook[title]
            column = _find_column(workbook, title, header)
            value = None
            for row in sheet.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS, values_only=True):
                label = _normalized_text(row[0])
                if JAPANESE_DATE.search(str(row[0] or "")) or label == "全国":
                    candidate = _number(row[column - 1])
                    if candidate is not None:
                        value = candidate
                        break
            if value is None:
                raise WorkbookFormatError(f"{title}: national occupancy rate missing")
            if not 0 <= float(value) <= 100:
                raise WorkbookFormatError(
                    f"{title}: invalid national occupancy rate: {value}"
                )
            records.append(
                NationalOccupancyRecord(
                    year=year,
                    month=month,
                    occupancy_rate=float(value),
                    release_type=release_type,
                )
            )
    finally:
        workbook.close()
    return records
